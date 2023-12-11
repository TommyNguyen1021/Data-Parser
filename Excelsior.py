import csv
from numpy import double
import openpyxl
import os
from openpyxl import load_workbook




def open_csv(filename, ws):
    
    with open(filename+".csv") as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            for i in range(len(row)):
                try:
                    row[i] = double(row[i])
                except:
                    None
            ws.append(row)



def open_in_excel(fn, test):
    filename = os.path.abspath(fn)
    wb = load_workbook(os.path.abspath("parsed_data/parsed_data_template" + '.xlsx'))
    #wb.create_sheet('test')
    #ws = wb["test"]
    #ws.sheet_state = 'hidden'
    #open_csv(filename, ws)  
    wb.save(filename+'.xlsx')
    #os.system(filename+'.xlsx')


#open_in_excel("parsed_data/vili_P9HK27_0j_07_P0021_085C_210908_read_shmoo")